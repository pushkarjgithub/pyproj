import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
# from openpyxl.styles.colors import GREEN

# refer https://gitlab.com/nanuchi/python-programming/-/tree/master/spreadsheet

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}
product_under_10_inv = {}

print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # calculate total number of product per supplier
    if supplier_name in product_per_supplier:
        product_per_supplier[supplier_name] = product_per_supplier.get(supplier_name) + 1
    else:
        product_per_supplier[supplier_name] = 1



    # calculate total product inventory value (unit * price) per supplier
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] = total_value_per_supplier.get(supplier_name) + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # find product number whose inventory is less than 10
    if inventory < 10:
        product_under_10_inv[int(product_num)] = int(inventory)

    # add new column for total inventory price for each row and create a new output file
    inventory_price.value = inventory * price

print(product_per_supplier)
print(total_value_per_supplier)
print(product_under_10_inv)

# print(product_list.cell(1, 1).value)
product_list.cell(1, 5).value = "Total Inventory Price"
product_list['E1'].font = Font(bold=True)
product_list['E1'].fill = PatternFill(start_color="FFC7CE", fill_type = "solid")
# product_list['E1'].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
# product_list['E1'].fill = PatternFill(fgColor=YELLOW, fill_type = "solid")
# print(product_list.cell(1, 5).value)

# creating a new output file
inv_file.save("spreadsheet1_inventory_output.xlsx")
